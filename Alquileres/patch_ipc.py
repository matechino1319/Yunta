import json
import openpyxl
from datetime import datetime

# Load data.json
with open('data.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Load Excel workbook
wb = openpyxl.load_workbook('sheet.xlsx', data_only=True)
if 'IPC' in wb.sheetnames:
    ws = wb['IPC']
    
    new_ipc_data = []
    
    # Iterate through rows in IPC sheet (assuming A=date, B=Nac, C=Cuyo)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            try:
                date_val = row[0]
                if isinstance(date_val, str):
                    # sometimes Excel dates are parsed as strings
                    pass
                elif isinstance(date_val, datetime):
                    mes_str = date_val.strftime('%Y-%m')
                    
                    nac_val = float(row[1]) if row[1] is not None else 0
                    cuy_val = float(row[2]) if row[2] is not None else nac_val
                    
                    new_ipc_data.append({
                        "mes": mes_str,
                        "nac": nac_val,
                        "cuy": cuy_val
                    })
            except Exception as e:
                pass
                
    if new_ipc_data:
        # Sort data
        new_ipc_data.sort(key=lambda x: x['mes'])
        data['IPC_DATA'] = new_ipc_data
        
        with open('data.json', 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print("Updated data.json with IPC data from Excel. Extracted", len(new_ipc_data), "records.")
    else:
        print("Could not extract any valid IPC data from Excel.")
else:
    print("IPC sheet not found in Excel.")
