import openpyxl
import json

wb = openpyxl.load_workbook('sheet.xlsx', data_only=False)

formulas = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    formulas[sheet_name] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                formulas[sheet_name].append({
                    'cell': cell.coordinate,
                    'formula': str(cell.value)
                })

with open("formulas_output.json", "w") as f:
    json.dump(formulas, f, indent=2)

print("Done")
